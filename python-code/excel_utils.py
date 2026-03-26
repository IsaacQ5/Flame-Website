import openpyxl as excelSheet
from openpyxl.styles import PatternFill

# A solid black fill pattern used to visually separate sections in the Excel output.
BLACKFILL = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")


class GenerateExcel:
    def __init__(self, excel_values, list_of_videos=None, workbook=None):
        # The values from the standoff-distances 
        self.excel_values = excel_values
        # Optional list of filenames used to label columns when batch processing.
        self.list_of_videos = list_of_videos
        # Optional flag to include zero measurements (unused by default).
        self.COUNT_ZERO_MEASURE = False

        # Per-run workbook and sheet (avoid global state).
        self.workbook = workbook or excelSheet.Workbook()
        self.sheet = self.workbook.active

    def excel(self, video=None):
        """Formats the Excel file with summary statistics and displacement."""

        # each column's meaning in the excel file, used for formatting the output
        excel_indexes = {
            "distance": 0,
            "distance_avg": 1,
            "median": 2,
            "displacement": 3,
            "displacement_avg": 4,
            "blackfill": 5,
        }

        # Column offset for each video block (6 columns per block).
        number_for_excel = 0
        for i in range(1, (len(self.excel_values) + 1)):  # loops for each video measured

            # gets stats from the values
            values = self.excel_values["Distance" + str(i)]

            # Removes 0 values
            values_without_zero = [value for value in values if value != 0]
            
            # to avoid dividing by zero when calculating the average
            if values_without_zero == []:
                values_without_zero = [0]
                
            # calculates the average, median, and displacement for the distance values
            distance_average = self.getDistanceAverage(values_without_zero)
            median = self.getMedian(values_without_zero)
            displacement = self.getDisplacement(values)
            displacement_average = self.getDisplacementAverage(displacement)

            # Checks if its a list of videos or just one.
            if type(video) == str:
                # name of the test
                self.sheet[f"{self.letterValue(number_for_excel)}1"] = f"Test {video}"
            else:
                # name of the test
                video_name = self.list_of_videos[i - 1]
                self.sheet[f"{self.letterValue(number_for_excel)}1"] = video_name[
                    video_name.index("Test") :
                ]

            # Loops for each value measured.
            for j in range(2, len(values) + 1):
                # makes cells for the distance values
                self.sheet[f"{self.letterValue(number_for_excel)}{j}"] = values[j - 2]
                # makes cells for the displacement values
                self.sheet[
                    f"{self.letterValue(number_for_excel + excel_indexes['displacement'])}{j}"
                ] = displacement[j - 2]
                # makes cell of a black fill for the serpration
                self.sheet[
                    f"{self.letterValue(number_for_excel + excel_indexes['blackfill'])}{j-1}"
                ].fill = BLACKFILL

            # Makes the cells for distance average.
            self.sheet[
                f"{(self.letterValue(number_for_excel+excel_indexes['distance_avg']))}1"
            ] = "AVG"
            self.sheet[
                f"{(self.letterValue(number_for_excel+excel_indexes['distance_avg']))}2"
            ] = distance_average

            # Makes the cells for the median.
            self.sheet[
                f"{(self.letterValue(number_for_excel+excel_indexes['median']))}1"
            ] = "Median"
            self.sheet[
                f"{(self.letterValue(number_for_excel+excel_indexes['median']))}2"
            ] = median

            # Makes the cells for the displacement.
            self.sheet[
                f"{(self.letterValue(number_for_excel+excel_indexes['displacement']))}1"
            ] = "Displacement"
            self.sheet.column_dimensions[
                f"{(self.letterValue(number_for_excel+excel_indexes['displacement']))}"
            ].width = 12

            # Makes the cells for displacement average.
            self.sheet[
                f"{(self.letterValue(number_for_excel+excel_indexes['displacement_avg']))}1"
            ] = "AVG"
            self.sheet[
                f"{self.letterValue(number_for_excel+excel_indexes['displacement_avg'])}2"
            ] = displacement_average

            # Formats the cells based on the size of each section.
            number_for_excel += len(excel_indexes)

        # Return the workbook so callers can decide where to save it.
        return self.workbook

    ###helper functions for calculating the average, median, and displacement for the distance values
    def getDistanceAverage(self, values):
        """Gets the average of distance values."""
        return sum(values) / len(values)

    def getMedian(self, values):
        """Gets the median for distance values."""
        sorted_values = sorted(values)
        return sorted_values[len(sorted_values) // 2]

    def getDisplacement(self, values):
        """Gets the displacement of the distances."""
        displacements = []
        for i in range(1, len(values)):
            displacements.append(abs(values[i] - values[i - 1]))

        return displacements

    def getDisplacementAverage(self, displacement):
        """Gets the displacement average."""
        # A single measured frame produces no displacement values, so treat
        # that case as a zero average instead of dividing by zero.
        if not displacement:
            return 0
        return sum(displacement) / len(displacement)

    def letterValue(self, i):
        """Maps a number to its corresponding Excel column letter(s)."""
        if i < 26:
            return chr(ord("A") + i)
        return chr(ord("A") + (i // 26) - 1) + chr(ord("A") + (i % 26))
